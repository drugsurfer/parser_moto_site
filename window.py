from tkinter import *
import pandas as pd
import threading
import requests
import bs4
import openpyxl as xl
import numpy as np

class Parser:
    def __init__(self, table_path: str):
        self.table_path = table_path # путь к таблице
        self.link = {'equipment': 'https://pitstopmoto.ge/en/store/equipment/',
                    'parts': 'https://pitstopmoto.ge/en/store/parts-and-accessories/'} # хранит урлы для перехода на ключевые страницы


    def go_to_tab(self, name_tab: str) -> str:
        '''
        возвращает 1 страницу с элементами из нужной области
        '''
        return requests.get(self.link[name_tab]).text


    def go_to_tab_with_link(self, link: str) -> str:
        '''
        Переходит по нужной ссылке и возвращает исходный код этой страницы
        '''
        return requests.get(link).text


    def go_link_for_next_page(self, soup: object, url: str, flag: bool = False):
        '''
        Переходит на следующую страницу с элементами
        '''
        url = url.split('?')[0]
        if flag:
            link =  url + soup.find_all('div', class_='col-md-12 col-xs-12 MID mt-3 TC')[0].find_all('a')[8].get('href')
        else:
            link =  url + soup.find_all('div', class_='col-md-12 col-xs-12 MID mt-3 TC')[0].find_all('a')[-2].get('href')
        return requests.get(link).text, link

    def get_cnt_pages(self, soup: object) -> int:
        '''
        Возвращает общее кол-во страниц
        по которым необходимо пройтись в текущем разделе
        '''
        return int(soup.find_all('div', class_ = 'col-md-12 col-xs-12 MID mt-3 TC')[0].find_all('a', class_='PG USR')[-1].text)

    def get_item_lst(self, soup: object) -> list:
        '''
        возвращает массив с ссылками на айтемы
        html: объект bs4
        return: список с ссылками на страницы с подробным описанием
        '''
        items = []
        table = soup.find_all('div', class_='col-md-3 col-6') # таблица с нужными нам элементами
        for item in table:
            items.append('https://pitstopmoto.ge' + item.find('a', class_='h-100 TC').get('href'))
        return items

    def get_size_item(self, text: str) -> list:
        '''
        Метод парсит описание и ищет в тексте размеры
        return : список с размерами, если список пуст или [''], то размеров нет в описании к товару
        '''
        sizes_dict = [('XXS', '2XS'), ('XS', '1XS'), ('S', 'S'), ('M', 'M'), ('L', 'L'), ('XL', '1XL'), ('XXL', '2XL'), 
                        ('XXXL', '3XL'), ('XXXXL', '4XL'), ('XXXXXL', '5XL')]
        sizes = []
        s = text[text.find('Sizes') : text.find('Sizes') + 60]
        if len(s) == 0:
            s = text[text.find('sizes') : text.find('sizes') + 60]
        tmp = s[s.find('(') + 1 : s.find(')')].split('|')
        if len(tmp) == 1 and tmp[0] != '':
            if 'all long sizes' in tmp[0].lower():
                tmp = ['all long sizes']
            elif 'to' in tmp[0]:
                start, end = -1, 0
                for i, sym in enumerate(sizes_dict):
                    if sym[0] in tmp[0]:
                        if start == -1:
                            start = i
                        else:
                            end = i
                    elif sym[1] in tmp[0]:
                        if start == -1:
                            start = i
                        else:
                            end = i
                if start == -1 or end == 0:
                    tmp = ['']
                else:
                    tmp = [sizes_dict[i][0] for i in range(start, end + 1)]
            elif '-' in tmp[0]:
                start, end = -1, 0
                for i, sym in enumerate(sizes_dict):
                    if sym[0] in tmp[0]:
                        if start == -1:
                            start = i
                        else:
                            end = i
                    if sym[1] in tmp[0]:
                        if start == -1:
                            start = i
                        else:
                            end = i
                if start == -1 or end == 0:
                    tmp = ['']
                else:
                    tmp = [sizes_dict[i][0] for i in range(start, end + 1)]
            else:
                tmp = ['']
        for i in tmp:
            for j in i.split('-'):
                sizes.append(j)
        return sizes

    def get_info_about_item(self, soup: object) -> dict:
        '''
        Парсим информацию об айтеме 
        Инфа:
        1) Title +
        2) Category +
        3) Brand +
        4) Size +
        5) Price +
        6) Description +
        7) Material - есть только у нескольких товаров в названии
        8) Image +
        9) SKU +
        10) parentSKU +
        '''
        info = {}
        tmp = soup.find_all('div', class_='row mx-0 w-100 mobd-none')
        if len(tmp) == 0: # у некоторых товаров страница пустая, хоть они и есть на странице как айтем, эта проверка для этого
            return None
        main_title = tmp[0]
        info['Title'] = main_title.find('h1').text # Получаем Title
        info['Category'] = soup.find_all('div', class_='col-md-12 BRD mt-2')[0].find_all('a')[2].text # Получаем Category
        info['Brand'] = main_title.find('a', class_='IB').get('href').split('=')[1] # Получаем brand
        info['Price'] = soup.find_all('div', class_='PBP')[0].find('span').text # получаем price
        info['Description'] = soup.find_all('div', class_='col-md-12 DETAILS')[0].text # получаем description
        tmp = soup.find_all('div', class_='col-md-12 ZW text-center')[0].find('img')
        if tmp is not None:
            info['Image'] = 'https://pitstopmoto.ge' + soup.find_all('div', class_='col-md-12 ZW text-center')[0].find('img').get('src')
        else:
            info['Image'] = ''
        info['Size'] = self.get_size_item(info['Description'])
        info['ParentSKU'] = soup.find_all('div', class_ = 'CDE')[0].text # предусмотреть изменение
        info['SKU'] = soup.find_all('div', class_ = 'CDE')[0].text
        info['in_storage'] = 0
        return info


class ExcelParser:
    def __init__(self, table_path: str, sheet_name: str):
        self.table_path = table_path

        if self.check_sheet(sheet_name):
            self.sheet_name = sheet_name
        else:
            self.create_sheet(sheet_name)
            self.sheet_name = sheet_name

        self.df = None

    def reverse_flag(self):
        '''
        Проходит по таблице
        если в столбце 'flag' стоит 1, то мы заменяем ее на 0
        если стоит 0 и в столбце 'in_storage' стоит 0, то удаляем иначе не трогаем
        '''
        del_index = []
        for index, data in self.df.iterrows():
            row = data.values
            if row[self.column_name['flag']] == 1:
                row[self.column_name['flag']] = 0
            elif (row[self.column_name['flag']] == 0) and row[self.column_name['in_storage']] == 1:
                del_index.append(index)
        self.df.drop(del_index, axis=0)


    def check_sheet(self, sheet_name: str) -> bool:
        '''
        Проверяет наличие листа в ексельке
        '''
        try:
            book = xl.load_workbook(self.table_path)
            ws = book[sheet_name]
        except KeyError:
            return False
        return True

    def create_sheet(self, sheet_name: str):
        '''
        создает лист в ексельке
        '''
        book = xl.load_workbook(self.table_path)
        book.create_sheet(sheet_name)

    def convert_to_df(self) -> bool:
        '''
        Конвертиирует в pandas таблицу
        '''
        book = xl.load_workbook(self.table_path)
        try:
            sheet = book[self.sheet_name]
        except Exception:
            return False
        self.df = pd.DataFrame(sheet.values)
        self.column_name = self.get_column_name()
        return True

    def get_column_name(self) -> dict:
        '''
        Возвращает словарь для связи словаря который был взят с сайта
        со строками в таблице
        dict: {key - название столбца, value - номер столбца в pandas таблице}
        '''
        column_name = {}
        for i, name in enumerate(self.df.iloc[0].values):
            if isinstance(name, str):
                column_name[name] = i
        return column_name

    def check_in_storage(self, info: list) -> bool:
        '''
        Проверяет наличие товара в таблице на складе, т.е. не трогаем элемент
        '''
        column_name = self.column_name
        '''
        if self.df.iloc[index].values[column_name['in_storage']] == '1':
            return True
        return False
        '''
        if self.df.loc[(self.df[0] == info[0]) & (self.df[1] == info[1])].values[0][column_name['in_storage']] == '1':
            return True
        return False

    def check_in_element(self, element_info: dict):
        '''
        Возвращает есть элемент с такими данными в таблице или нет и его индекс
        '''
        column_name = self.column_name
        for col_name, data in self.df.iterrows():
            if col_name == 0:
                continue
            if element_info['Title'] == data[column_name['Title']]:
                if element_info['ParentSKU'] == data[column_name['ParentSKU']]:
                    return True, col_name
        return False, None

    def add_item(self, element_info: dict, number: int):
        '''
        Добавляет в таблицу новый элемент
        '''
        column_name = self.column_name
        add_values = [None] * len(column_name)
        for elem in element_info.keys():
            if elem == 'Size':
                if isinstance(element_info[elem], list):
                    if len(element_info[elem]) > 0:
                        cnt = 1
                        for s in element_info[elem]:
                            info = element_info.copy()
                            info['Size'] = s
                            self.add_item(info, cnt)
                            cnt += 1
                        return
                    else:
                        add_values[column_name[elem]] = ''
                else:
                    add_values[column_name[elem]] = element_info[elem]
            if elem == 'SKU':
                if isinstance(element_info['Size'], str):
                    add_values[column_name[elem]] = element_info[elem] + '-' + str(number)
            else:
                add_values[column_name[elem]] = element_info[elem]
        add_values[-1] = 1 # добавляем флаг, 1 - мы с ним работали
        self.df.loc[len(self.df.index)] = add_values
        
    def replace_item(self, element_info: dict, number: int):
        '''
        Заменяет в таблице элемент по индексу
        '''
        column_name = self.column_name
        replace_values = [None] * len(column_name)
        for elem in element_info.keys():
            if elem == 'Size':
                '''
                if element_info[elem] == [] or element_info[elem] == ['']:
                    replace_values[column_name[elem]] = ''
                else:
                    replace_values[column_name[elem]] = ' '.join(element_info[elem])
                '''
                if isinstance(element_info[elem], list):
                    if len(element_info[elem]) > 0:
                        cnt = 0
                        for s in element_info[elem]:
                            info = element_info.copy()
                            info['Size'] = s
                            self.replace_item(info, cnt + 1)
                        return
                    else:
                        replace_values[column_name[elem]] = ''
                else:
                    replace_values[column_name[elem]] = element_info[elem]
            if elem == 'SKU':
                if isinstance(element_info['Size'], str):
                    replace_values[column_name[elem]] = element_info[elem] + '-' + str(number)
            else:
                replace_values[column_name[elem]] = element_info[elem]
        replace_values[-1] = 1 # добавляем флаг, 1 - мы с ним работали
        if not self.check_in_storage([replace_values[0], replace_values[1]]):
            self.df[np.logical_and(self.df.values == replace_values[0], self.df.values == replace_values[1])] = replace_values

class Main:
    def __init__(self, sheet_name: str, file_path: str, window: object):
        self.sheet_name = sheet_name
        self.url = 'https://pitstopmoto.ge/en' # хранит текущий url
        self.file_path = file_path # хранит путь к файлу
        self.parserObj = Parser(self.file_path)
        self.excel_parserObj = ExcelParser(self.file_path, sheet_name)
        self.windowObj = window

    def check_connection(self, url: str) -> bool:
        res = requests.get(url)
        if res.status_code == 200: return True
        else: return False

    def check_file(self, path: str) -> bool:
        try:
            book = xl.load_workbook(path)
        except Exception:
            return False
        return True

    def check_item_in_table(self, item_info: dict):
        '''
        Получает информацию об айтеме и проверет 
        его наличие в листе ексельки
        '''
        flag, index = self.excel_parserObj.check_in_element(item_info)
        if flag:
            return True, index
        return False, None

    def replace_item_in_table(self, item_info: dict, index: int):
        '''
        Заменяет элемент в таблице на айтем с сайта
        '''
        self.excel_parserObj.replace_item(item_info, index)

    def add_item_in_table(self, item_info: dict):
        '''
        Добавляет айтем в таблицу с данными
        '''
        self.excel_parserObj.add_item(item_info, 0)
        
    def main(self):
        if not self.check_connection(self.url):
            self.windowObj.print_log('Проблемы с подключением к сайту\n')
            return
        if not self.check_file(self.file_path):
            self.windowObj.print_log('Проблемы с файлом\n')
            return
        self.windowObj.print_log('Файл и подключение - ок\n')
        if not self.excel_parserObj.convert_to_df():
            self.windowObj.print_log('Ошибка в названии листов в excel файле.\
                Листы должны называться: equipment, parts\n')
            return
        # Начали парсить equipments
        self.url = self.parserObj.link[self.sheet_name]
        
        count_pages = self.parserObj.get_cnt_pages(bs4.BeautifulSoup(self.parserObj.go_to_tab_with_link(self.url), 'html.parser'))
        # Проход по страницам
        for i in range(count_pages):
            self.windowObj.print_log('Парсится {}-я страница раздела {}\n'.format(i, self.sheet_name))
            if i == 0:
                items = self.parserObj.get_item_lst(
                    bs4.BeautifulSoup(self.parserObj.go_to_tab(self.sheet_name), 'html.parser'))
            elif i == 4 and self.sheet_name == 'parts':
                self.url = self.url.replace('1', '4')
                req, link = self.parserObj.go_link_for_next_page(
                        bs4.BeautifulSoup(self.parserObj.go_to_tab_with_link(self.url), 'html.parser'), self.url)
                items = self.parserObj.get_item_lst(bs4.BeautifulSoup(req, 'html.parser'))
                self.url = link
            else:
                if i > 8:
                    req, link = self.parserObj.go_link_for_next_page(
                        bs4.BeautifulSoup(self.parserObj.go_to_tab_with_link(self.url), 'html.parser'), self.url, True)
                else:
                    req, link = self.parserObj.go_link_for_next_page(
                        bs4.BeautifulSoup(self.parserObj.go_to_tab_with_link(self.url), 'html.parser'), self.url)
                items = self.parserObj.get_item_lst(bs4.BeautifulSoup(req, 'html.parser'))
                self.url = link
        
            # Проходим по айтемам на странице  
            for item in items:
                soup = bs4.BeautifulSoup(self.parserObj.go_to_tab_with_link(item), 'html.parser')
                info_about_item = self.parserObj.get_info_about_item(soup) # словарь с инфой об айтеме
                if info_about_item is None:
                    continue
                in_table, index = self.check_item_in_table(info_about_item)
                if in_table:
                    self.replace_item_in_table(info_about_item, 0)
                else:
                    self.add_item_in_table(info_about_item)
        self.excel_parserObj.reverse_flag()
        self.windowObj.check_in_df(self.sheet_name, self.excel_parserObj.df)


class Window:
    def __init__(self, WIDTH, HEIGHT, title='Parser', resiazable=(False, False)):
        self.root = Tk()
        self.root.title(title)
        self.root.geometry(f"{WIDTH}x{HEIGHT}+200+200")
        self.root.resizable(resiazable[0], resiazable[1])
        self.font = ("Calibri", 15, "bold")

        self.entry_file_path = Entry(self.root, width=30)
        self.entry_file_path.insert(0, 'Путь к файлу')

        self.help = Text(self.root, width=30, height=10)
        self.help.insert(END, 'Путь к файлу может быть любой. В файле обязательно должны \
            быть два листа: equipment и parts. Столбцы в листах должны быть названы: \
                Title, SKU, ParentSKU, Category, Brand, Color, Size, Price,\
                    Description, Image, in_storage, flag. Столбцы и имена листов должны быть \
                        названы именно так(регистр важен).')

        self.button_start = Button(self.root, command=lambda: self.start(),
                                text='Запустить', width=30, height=5)

        self.log = Text(self.root, width=30, height=10)

        self.df1 = None
        self.df2 = None

    def run(self):
        self.draw_widgets()
        self.root.mainloop()

    def draw_widgets(self):
        self.entry_file_path.pack()
        self.help.pack()
        self.button_start.pack()
        self.log.pack()

    def print_log(self, text: str):
        self.log.insert(END, text)

    def start(self):
        if self.check():
            m = Main('equipment', self.entry_file_path.get(), self)
            thread = threading.Thread(target=m.main)
            thread.start()
            m = Main('parts', self.entry_file_path.get(), self)
            thread_ = threading.Thread(target=m.main)
            thread_.start()


    def check_in_df(self, sheet_name, df):
        if sheet_name == 'equipment':
            self.df1 = df
        elif sheet_name == 'parts':
            self.df2 = df
            self.write_excel()


    def write_excel(self):
        with pd.ExcelWriter(self.entry_file_path.get()) as writer:
                self.df1.to_excel(writer, sheet_name='equipment', index=False, header=False)
                self.df2.to_excel(writer, sheet_name='parts', index=False, header=False)

    def check(self) -> bool:
        text = self.entry_file_path.get()
        if text == 'Путь к файлу' or len(text) == 0:
            self.log.insert(END, 'Некорректный ввод пути к файлу\n')
            print('Некорректный ввод пути к файлу\n')
            return False
        m = Main('equipment', self.entry_file_path.get(), self)
        if not m.check_file(m.file_path):
            return False
        return True
        


window = Window(400, 350)
window.run()
    
